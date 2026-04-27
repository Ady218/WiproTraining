from openpyxl import Workbook, load_workbook
import os

def write_excel(filename):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Name', 'Age'])
    sheet.append(['Arpit', '22'])
    sheet.append(['Messi', '30'])
    workbook.save(filename)

def read_excel(filename):
    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True):
        print(f'Name: {row[0]}, Age: {row[1]}')

# Function to delete an Excel file
def delete_excel(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f'{filename} deleted successfully')
    else:
        print(f'{filename} does not exists')



filename = "data.xlsx"
write_excel(filename)
print('Data read from Excel file: ')
read_excel(filename)
write_excel(filename)